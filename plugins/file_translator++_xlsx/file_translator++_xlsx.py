from GalTransl import LOGGER
from GalTransl.GTPlugin import GFilePlugin
import openpyxl
import re


class file_plugin(GFilePlugin):
    def gtp_init(self, plugin_conf: dict, project_conf: dict):
        """
        This method is called when the plugin is loaded.在插件加载时被调用。
        :param plugin_conf: The settings for the plugin.插件yaml中所有设置的dict。
        :param project_conf: The settings for the project.项目yaml中common下设置的dict。
        """
        # 打印提示的方法，打印时请带上模块名，以便区分日志。
        self.pname = plugin_conf["Core"].get("Name", "")
        settings = plugin_conf["Settings"]
        LOGGER.debug(
            f"[{self.pname}] 当前配置：是否自动识别名称:{settings.get('是否自动识别名称', False)}")
        # 读取配置文件中的设置，并保存到变量中。
        self.是否自动识别名称 = settings.get("是否自动识别名称", True)
        self.名称识别正则表达式 = settings.get(
            "名称识别正则表达式", r"^(?P<name>.*?)「(?P<message>.*?)」$")

    def load_file(self, file_path: str) -> list:
        """
        This method is called to load a file.
        加载文件时被调用。
        :param file_path: The path of the file to load.文件路径。
        :return: A list of objects with message and name(optional).返回一个包含message和name(可空)的对象列表。
        """
        if not file_path.endswith(".xlsx"):
            # 检查不支持的文件类型并抛出TypeError
            raise TypeError("File type not supported.")

        json_list = self.read_xlsx_to_json(file_path)

        return json_list

    def save_file(self, file_path: str, transl_json: list):
        """
        This method is called to save a file. 保存文件时被调用。
        :param file_path: The path of the file to save. 保存文件路径。
        :param transl_json: A list of objects same as the return of load_file(). load_file提供的json在翻译message和name后的结果。
        :return: None.
        """
        EXCEL_TITLE_LIST = ["Original Text", "Initial",
                            "Machine translation", "Better translation", "Best translation"]

        LOGGER.debug(f"[{self.pname}] Saving file {file_path}")

        original_file_path_01 = file_path.replace("gt_output", "gt_input")
        original_file_path_02 = file_path.replace("gt_output", "json_input")
        # 读取原文件
        if original_file_path_01:
            original_file_path = original_file_path_01
        elif original_file_path_02:
            original_file_path = original_file_path_02
        else:
            original_file_path = ""

        try:
            # 读取原文件的Original Text
            original_texts = []
            if original_file_path:
                original_workbook = openpyxl.load_workbook(original_file_path)
                original_sheet = original_workbook.active
                original_texts = [cell.value for cell in original_sheet['A']]
                # 检查并跳过标题行
                if original_texts and original_texts[0] == "Original Text":
                    original_texts = original_texts[1:]

            # 创建一个新的工作簿
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            # 写入标题行
            for i, title in enumerate(EXCEL_TITLE_LIST):
                sheet.cell(row=1, column=i+1, value=title)

            # 重组翻译结果
            for i, transl_obj in enumerate(transl_json):
                if self.是否自动识别名称 and transl_obj["name"] is not None and transl_obj["name"]!= "":
                    translated_text = f"{transl_obj['name']}\n「{transl_obj['message']}」"
                else:
                    translated_text = transl_obj["message"]

                # 写入 Original Text 栏
                if i < len(original_texts):
                    sheet.cell(row=i+2, column=1, value=original_texts[i])
                else:
                    sheet.cell(row=i+2, column=1, value=translated_text)

                # 写入 Machine translation 栏
                sheet.cell(row=i+2, column=3, value=translated_text)

            workbook.save(file_path)
        except Exception as e:
            LOGGER.error(f"Error saving file {file_path}: {e}")
            raise e

    def gtp_final(self):
        """
        This method is called after all translations are done.
        在所有文件翻译完成之后的动作，例如输出提示信息。
        """
        pass

    def read_xlsx_to_json(self, file_path: str) -> list:
        """
        读取xlsx文件，返回一个包含message和name(可空)的对象列表。
        :param file_path: xlsx文件路径。
        :return: 包含message和name(可空)的对象列表。
        """
        json_list = []

        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            # 获取第一列的所有内容
            first_column_values = [cell.value for cell in sheet['A']]

            # 检查并跳过标题行
            if first_column_values and first_column_values[0] == "Original Text":
                first_column_values = first_column_values[1:]

            for row in first_column_values:
                if row is None:
                    json_list.append({"name": "", "message": ""})
                    continue

                name = ""
                message = row.strip()

                if self.是否自动识别名称:
                    match = re.search(self.名称识别正则表达式, row, re.DOTALL)
                    if match:
                        name = match.group("name").strip()
                        message = match.group("message").strip()

                json_list.append({"name": name, "message": message})

        except FileNotFoundError:
            LOGGER.error(f"File not found: {file_path}")
            raise
        except openpyxl.utils.exceptions.InvalidFileException:
            LOGGER.error(f"Invalid file format: {file_path}")
            raise
        except Exception as e:
            LOGGER.error(f"Error loading file {file_path}: {e}")
            raise

        return json_list
